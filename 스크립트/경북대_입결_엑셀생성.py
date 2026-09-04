# -*- coding: utf-8 -*-
import json, os, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

A=json.load(open('/tmp/knu_all.json'))
for _r in A:
    _r['수능최저']='적용' if _r.get('최저기준통과 인원') is not None else '미적용'
    if _r['수능최저']=='미적용':
        _r['최저기준통과 인원']='—'; _r['실질경쟁률(최저적용후)']='—'
OUT='/home/user/book/엑셀/경북대_농업생명_생태환경_6개모집단위_입결.xlsx'
YEARS=[2024,2025,2026]
DEPTS=['응용생명과학부','식물의학과','원예과학과','농업생명과학대학 자율학부','식물자원학과','곤충생명과학과']
COLLEGE={'응용생명과학부':'농업생명과학대학','식물의학과':'농업생명과학대학','원예과학과':'농업생명과학대학',
 '농업생명과학대학 자율학부':'농업생명과학대학','식물자원학과':'생태환경대학','곤충생명과학과':'생태환경대학'}

METRICS=['모집인원','지원인원','경쟁률','입학인원(등록)','추합최종번호','추합인원','수능최저',
 '최저기준통과 인원','실질경쟁률(최저적용후)','등급 평균','등급 표준편차','등급 50%컷','등급 70%컷','등급 85%컷',
 '반영점수 만점','반영점수 평균','반영점수 50%컷','반영점수 70%컷','반영점수 85%컷']

NAVY='1F3864'; GREEN='1F5C3A'; LGREEN='E2EFDA'; LGRAY='F2F2F2'; GOLD='FFF2CC'
thin=Side(style='thin',color='BFBFBF')
BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
HF=Font(name='맑은 고딕',size=10,bold=True,color='FFFFFF')
BF=Font(name='맑은 고딕',size=10)
BFB=Font(name='맑은 고딕',size=10,bold=True)
TF=Font(name='맑은 고딕',size=14,bold=True,color=NAVY)
CEN=Alignment(horizontal='center',vertical='center',wrap_text=True)
LEF=Alignment(horizontal='left',vertical='center',wrap_text=True)

wb=openpyxl.Workbook(); wb.remove(wb.active)

def style_table(ws,r0,nrow,ncol,hdr_rows=1):
    for r in range(r0,r0+hdr_rows):
        for c in range(1,ncol+1):
            cell=ws.cell(r,c); cell.font=HF; cell.alignment=CEN; cell.border=BOX
            cell.fill=PatternFill('solid',fgColor=GREEN)
    for r in range(r0+hdr_rows,r0+hdr_rows+nrow):
        for c in range(1,ncol+1):
            cell=ws.cell(r,c); cell.font=BF; cell.border=BOX
            cell.alignment=CEN if c>2 else LEF
            if (r-r0-hdr_rows)%2: cell.fill=PatternFill('solid',fgColor=LGRAY)
    ws.freeze_panes=ws.cell(r0+hdr_rows,1)

def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w

def title(ws,txt,sub=None,ncol=10):
    ws['A1']=txt; ws['A1'].font=TF
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    ws['A1'].alignment=LEF; ws.row_dimensions[1].height=24
    if sub:
        ws['A2']=sub; ws['A2'].font=Font(name='맑은 고딕',size=9,color='808080')
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
        ws['A2'].alignment=LEF
    return 4 if sub else 3

def dkey(d): return DEPTS.index(d)

# ---------------- 1. 안내 ----------------
ws=wb.create_sheet('①안내·자료범위')
ws['A1']='경북대학교 농업생명과학대학·생태환경대학 6개 모집단위 수시 입시결과'
ws['A1'].font=Font(name='맑은 고딕',size=16,bold=True,color=NAVY)
ws.merge_cells('A1:F1')
notes=[
 ('','',''),
 ('■ 대상 모집단위','',''),
 ('농업생명과학대학','응용생명과학부 / 식물의학과 / 원예과학과 / 농업생명과학대학 자율학부',''),
 ('생태환경대학(상주)','식물자원학과 / 곤충생명과학과',''),
 ('','',''),
 ('■ 자료 범위 — 요청은 6개년이나, 경북대 입학처가 공개하는 연도별 수치 자료는 3개년입니다','',''),
 ('학년도','공개 형태','비고'),
 ('2026학년도','연도별 수치표 (엑셀)','전 전형·전 지표 수록 — 본 파일에 반영'),
 ('2025학년도','연도별 수치표 (엑셀)','전 전형·전 지표 수록 — 본 파일에 반영'),
 ('2024학년도','연도별 수치표 (엑셀)','전 전형·전 지표 수록 — 본 파일에 반영'),
 ('2023학년도','분포표(산점도) PDF만','수치표 미공개. 합격자 평균·분포는 「⑧2023 분포표」 시트의 원문 이미지로 확인'),
 ('2022학년도','미공개','경북대 입학처 자료실/공지 어디에도 게시물 없음'),
 ('2021학년도','미공개','상동. 단, 특별전형(농어촌 등)은 4개년 합산본만 존재 → 「⑨참고_4개년 집계」'),
 ('','',''),
 ('■ 출처 (모두 경북대학교 입학처 공식 게시물)','',''),
 ('2026학년도 대학입학전형 입학 전형결과(수시/정시) 안내','ipsi1.knu.ac.kr › 자료실',''),
 ('2025학년도 대학입학전형 입학 전형결과(수시/정시) 안내','ipsi1.knu.ac.kr › 자료실',''),
 ('2024학년도 대학입학전형 입학 전형결과 안내','ipsi1.knu.ac.kr › 자료실',''),
 ('2023학년도 대입 결과 분포표','ipsi1.knu.ac.kr › 자료실',''),
 ('4개년(2023~2026학년도) 입결성적 자료 안내','ipsi1.knu.ac.kr › 자료실 (특별전형 한정, 연도 합산)',''),
 ('','',''),
 ('■ 용어 정의 (경북대 공시 기준)','',''),
 ('모집인원','최초 모집인원',''),
 ('경쟁률','지원인원 ÷ 모집인원',''),
 ('최저기준통과 인원','수능 최저학력기준을 충족한 지원자 수',''),
 ('실질경쟁률(최저적용후)','최저기준통과 인원 ÷ 모집인원  ← 요청하신 「최저 적용 후 경쟁률」',''),
 ('추합최종번호','최종 예비합격 순번 (충원 규모의 상한 지표)',''),
 ('추합인원','실제 추가합격 인원',''),
 ('등급 평균','입학자(등록자) 학생부 교과등급 평균  ← 요청하신 「평균」',''),
 ('등급 50%컷','입학자 등급 상위 50% 지점  ← 「최초합」에 가장 가까운 공개 지표',''),
 ('등급 70%컷','입학자 등급 상위 70% 지점  ← 일반적인 지원 기준선',''),
 ('등급 85%컷','입학자 등급 상위 85% 지점  ← 「최저(합격선 하단)」에 가장 가까운 공개 지표',''),
 ('반영점수 ○○','교과전형의 환산총점(만점 500점) 기준 동일 지표',''),
 ('','',''),
 ('■ 수능 최저학력기준 적용 여부 (원문 표에 최저기준통과 인원이 기재된 전형 = 적용)','',''),
 ('학생부교과 교과우수자(일반)','적용','실질경쟁률 산출 가능'),
 ('학생부교과 지역인재','적용','실질경쟁률 산출 가능'),
 ('학생부종합 일반학생','적용','실질경쟁률 산출 가능'),
 ('학생부종합 지역인재','미적용','해당 6개 모집단위 기준. 원문 공란 → 본 파일에는 「—」'),
 ('학생부종합 영농창업인재','미적용','원문에 해당 열 자체가 없음 → 본 파일에는 「—」'),
 ('','',''),
 ('※ 경북대는 「최초합격자 성적」과 「최종 합격자 최저점」을 별도 공시하지 않습니다.','',''),
 ('   따라서 최초합 ≈ 50%컷, 최저 ≈ 85%컷으로 읽는 것이 실무 기준입니다.','',''),
 ('※ 2024학년도 원자료의 경쟁률 표기(예: "11.6:1")는 숫자만 남겨 정규화했습니다.','',''),
]
for i,(a,b,c) in enumerate(notes,start=2):
    ws.cell(i,1,a); ws.cell(i,2,b); ws.cell(i,3,c)
    for j in (1,2,3):
        cell=ws.cell(i,j); cell.font=BF; cell.alignment=LEF
    if a.startswith('■'):
        ws.cell(i,1).font=Font(name='맑은 고딕',size=11,bold=True,color=GREEN)
    if a in ('학년도',):
        for j in (1,2,3):
            cell=ws.cell(i,j); cell.font=HF; cell.fill=PatternFill('solid',fgColor=GREEN); cell.alignment=CEN
widths(ws,[46,52,62]); ws.sheet_view.showGridLines=False

# ---------------- 2. 종합표 ----------------
HDR=['학년도','단과대학','모집단위','전형구분','세부전형']+METRICS
def sortkey(r):
    order={'일반학생':0,'지역인재':1,'영농창업인재':2,'일반(교과우수자)':0,'사회통합(참고)':3,'농어촌(참고)':4,'일반(참고)':5,'SW특별(참고)':6}
    return (dkey(r['모집단위']), 0 if r['전형구분']=='학생부종합' else 1, order.get(r['세부전형'],9), r['학년도'])

ws=wb.create_sheet('②종합표(전체)')
r0=title(ws,'종합표 — 6개 모집단위 × 3개년 × 전 전형','모집단위 → 전형 → 학년도 순 정렬. 회색 음영은 참고 전형(사회통합·농어촌·논술·SW).',len(HDR))
for j,h in enumerate(HDR,1): ws.cell(r0,j,h)
rows=sorted(A,key=sortkey)
for i,r in enumerate(rows,start=r0+1):
    ws.cell(i,1,r['학년도']); ws.cell(i,2,r['단과대학']); ws.cell(i,3,r['모집단위'])
    ws.cell(i,4,r['전형구분']); ws.cell(i,5,r['세부전형'])
    for j,m in enumerate(METRICS,start=6): ws.cell(i,j,r.get(m))
style_table(ws,r0,len(rows),len(HDR))
for i,r in enumerate(rows,start=r0+1):
    if '참고' in r['세부전형']:
        for j in range(1,len(HDR)+1):
            ws.cell(i,j).fill=PatternFill('solid',fgColor='EDEDED')
            ws.cell(i,j).font=Font(name='맑은 고딕',size=10,color='7F7F7F')
    else:
        for j in (3,):
            ws.cell(i,j).font=BFB
widths(ws,[8,16,20,12,16]+[9]*6+[13,16]+[9,11,10,10,10]+[11,12,12,12,12])
ws.auto_filter.ref=f"A{r0}:{get_column_letter(len(HDR))}{r0+len(rows)}"

# ---------------- 3~7. 전형별 시트 ----------------
SHEETS=[
 ('③학종_일반학생','학생부종합','일반학생','학생부종합 일반학생전형 (모집단위별 3개년)'),
 ('④학종_지역인재','학생부종합','지역인재','학생부종합 지역인재전형 (모집단위별 3개년)'),
 ('⑤교과_일반','학생부교과','일반(교과우수자)','학생부교과 교과우수자전형 = 일반 (모집단위별 3개년)'),
 ('⑥교과_지역인재','학생부교과','지역인재','학생부교과 지역인재전형 (모집단위별 3개년)'),
 ('⑦학종_영농창업인재','학생부종합','영농창업인재','학생부종합 영농창업인재전형 — 원예과학과 (모집단위별 3개년)'),
]
SUBHDR=['모집단위','학년도']+METRICS
for name,cat,sub,ttl in SHEETS:
    ws=wb.create_sheet(name)
    sel=[r for r in A if r['전형구분']==cat and r['세부전형']==sub]
    sel.sort(key=lambda r:(dkey(r['모집단위']),r['학년도']))
    if sub in ('영농창업인재','지역인재') and cat=='학생부종합':
        note='이 전형은 수능 최저학력기준을 적용하지 않습니다 → 최저기준통과 인원·실질경쟁률은 「—」(원문 공란)'
    else:
        note='실질경쟁률(최저 적용 후 경쟁률) = 최저기준통과 인원 ÷ 모집인원'
    r0=title(ws,ttl,note,len(SUBHDR))
    for j,h in enumerate(SUBHDR,1): ws.cell(r0,j,h)
    i=r0
    prev=None
    for r in sel:
        i+=1
        ws.cell(i,1,r['모집단위']); ws.cell(i,2,r['학년도'])
        for j,m in enumerate(METRICS,start=3): ws.cell(i,j,r.get(m))
    n=len(sel)
    style_table(ws,r0,n,len(SUBHDR))
    # 모집단위 그룹 구분선 + 굵게
    i=r0; prev=None
    for r in sel:
        i+=1
        if r['모집단위']!=prev:
            for j in range(1,len(SUBHDR)+1):
                c=ws.cell(i,j); c.border=Border(left=thin,right=thin,bottom=thin,top=Side(style='medium',color=GREEN))
            ws.cell(i,1).font=BFB
            prev=r['모집단위']
        else:
            ws.cell(i,1,'')
    widths(ws,[20,8]+[9]*6+[13,16]+[9,11,10,10,10]+[11,12,12,12,12])
    if not sel:
        ws.cell(r0+1,1,'해당 전형으로 선발한 이력 없음').font=Font(name='맑은 고딕',size=10,italic=True,color='C00000')

# ---------------- 8. 연도비교 피벗 ----------------
PIV=['모집인원','경쟁률','실질경쟁률(최저적용후)','등급 평균','등급 50%컷','등급 70%컷','등급 85%컷']
PSHORT=['모집','경쟁률','실질경쟁률','평균','50%','70%','85%']
ws=wb.create_sheet('⑧연도비교(핵심지표)')
ncol=2+len(YEARS)*len(PIV)
r0=title(ws,'연도비교 — 모집단위 × 전형 × 3개년 핵심지표',
 '평균=입학자 등급평균 · 50%컷≈최초합 · 85%컷≈합격선 하단(최저) · 실질경쟁률=수능최저 통과 후 경쟁률',ncol)
ws.cell(r0,1,'모집단위'); ws.cell(r0,2,'전형')
ws.merge_cells(start_row=r0,start_column=1,end_row=r0+1,end_column=1)
ws.merge_cells(start_row=r0,start_column=2,end_row=r0+1,end_column=2)
c=3
for y in YEARS:
    ws.cell(r0,c,f'{y}학년도')
    ws.merge_cells(start_row=r0,start_column=c,end_row=r0,end_column=c+len(PIV)-1)
    for k,s in enumerate(PSHORT): ws.cell(r0+1,c+k,s)
    c+=len(PIV)
COMBOS=[('학생부종합','일반학생','학종 일반'),('학생부종합','지역인재','학종 지역인재'),
 ('학생부종합','영농창업인재','학종 영농창업'),('학생부교과','일반(교과우수자)','교과 일반(교과우수자)'),
 ('학생부교과','지역인재','교과 지역인재')]
idx={(r['학년도'],r['모집단위'],r['전형구분'],r['세부전형']):r for r in A}
i=r0+1; n=0; starts=[]
for d in DEPTS:
    first=True
    for cat,sub,lab in COMBOS:
        if not any((y,d,cat,sub) in idx for y in YEARS): continue
        i+=1; n+=1
        if first: starts.append(i); first=False
        ws.cell(i,1,d if len(starts)and ws.cell(i,1).value is None else d)
        ws.cell(i,2,lab)
        c=3
        for y in YEARS:
            r=idx.get((y,d,cat,sub))
            for k,m in enumerate(PIV): ws.cell(i,c+k, r.get(m) if r else None)
            c+=len(PIV)
style_table(ws,r0,n,ncol,hdr_rows=2)
# 연도 블록 색 구분 + 모집단위 병합느낌
for rr in range(r0+2,r0+2+n):
    c=3
    for bi,y in enumerate(YEARS):
        for k in range(len(PIV)):
            cell=ws.cell(rr,c+k)
            if bi%2==0: cell.fill=PatternFill('solid',fgColor='F7F9F4')
            else: cell.fill=PatternFill('solid',fgColor=LGREEN)
            if PSHORT[k]=='70%': cell.font=BFB
        c+=len(PIV)
    ws.cell(rr,1).font=BFB
for s in starts:
    for j in range(1,ncol+1):
        cell=ws.cell(s,j); b=cell.border
        cell.border=Border(left=b.left,right=b.right,bottom=b.bottom,top=Side(style='medium',color=GREEN))
prev=None
for rr in range(r0+2,r0+2+n):
    v=ws.cell(rr,1).value
    if v==prev: ws.cell(rr,1,'')
    else: prev=v
widths(ws,[20,20]+[7,8,9,7,7,7,7]*3)
ws.freeze_panes=ws.cell(r0+2,3)

# ---------------- 9. 참고 기타전형 ----------------
ws=wb.create_sheet('⑨참고_기타전형')
sel=[r for r in A if '참고' in r['세부전형']]
sel.sort(key=sortkey)
r0=title(ws,'참고 — 사회통합·농어촌·논술(AAT)·SW특별 전형 3개년',
 '요청 범위(학종 일반/지역인재, 교과 일반/지역인재, 영농창업) 밖이지만 같은 모집단위의 공시 자료라 함께 수록합니다.',len(HDR))
for j,h in enumerate(HDR,1): ws.cell(r0,j,h)
for i,r in enumerate(sel,start=r0+1):
    ws.cell(i,1,r['학년도']); ws.cell(i,2,r['단과대학']); ws.cell(i,3,r['모집단위'])
    ws.cell(i,4,r['전형구분']); ws.cell(i,5,r['세부전형'])
    for j,m in enumerate(METRICS,start=6): ws.cell(i,j,r.get(m))
style_table(ws,r0,len(sel),len(HDR))
widths(ws,[8,16,20,12,16]+[9]*6+[13,16]+[9,11,10,10,10]+[11,12,12,12,12])
ws.auto_filter.ref=f"A{r0}:{get_column_letter(len(HDR))}{r0+len(sel)}"

# ---------------- 10. 4개년 특별전형 집계 ----------------
import openpyxl as ox
src=ox.load_workbook('/home/user/book/입시결과/경북대_4개년_2023-2026_입결성적.xlsx',data_only=True)
ws=wb.create_sheet('⑩참고_4개년집계')
r0=title(ws,'참고 — 4개년(2023~2026학년도) 합산 입결 · 특별전형',
 '경북대가 별도 공개하는 유일한 4개년 자료. 연도별이 아닌 4년 합산치이며 특별전형(사회통합·농어촌·기초·사회배려·장애인)에 한정됩니다.',8)
H2=['전형명','단과대학','모집단위','4년 합산 모집인원','4년 합산 지원자','4년 합산 등록자','교과등급 평균','교과등급 85%']
for j,h in enumerate(H2,1): ws.cell(r0,j,h)
i=r0; cnt=0
for sname in src.sheetnames:
    s=src[sname]
    for row in s.iter_rows(values_only=True):
        if len(row)<8: continue
        mo=re.sub(r'\s+','',str(row[2] or ''))
        if mo not in [re.sub(r'\s+','',d) for d in DEPTS]: continue
        i+=1; cnt+=1
        ws.cell(i,1,sname); ws.cell(i,2,row[1]); ws.cell(i,3,str(row[2]).strip())
        for j,v in enumerate(row[3:8],start=4):
            try: ws.cell(i,j,float(v) if v is not None and str(v).strip() not in ('','-') else None)
            except: ws.cell(i,j,v)
style_table(ws,r0,cnt,len(H2))
widths(ws,[34,18,20,16,16,14,13,13])
if cnt==0: ws.cell(r0+1,1,'해당 모집단위 없음')

# ---------------- 11. 2023 분포표 이미지 ----------------
ws=wb.create_sheet('⑪2023 분포표')
r0=title(ws,'2023학년도 — 대입 결과 분포표 (원문 이미지)',
 '경북대는 2023학년도에 한해 수치표 대신 분포표(산점도)만 공개했습니다. ○=합격자, ×=불합격(수능최저 통과), ●=합격자 평균. 가로축은 학생부 교과등급.',8)
ws.sheet_view.showGridLines=False
caps=[('p3','학생부종합 일반학생전형 — 농업생명과학대학 (응용생명과학부·식물의학과)'),
      ('p5','학생부종합 일반학생전형 — 생태환경대학 (곤충생명과학과)  ※식물자원학과는 이 전형 선발 없음'),
      ('p6','학생부종합 지역인재전형 — 농업생명과학대학 (응용생명과학부·식물의학과)'),
      ('p7','학생부종합 농어촌학생전형 — 농업생명과학대학 (응용생명과학부)')]
row=r0
from PIL import Image as PILImage
os.makedirs('/tmp/k23x',exist_ok=True)
for key,cap in caps:
    ws.cell(row,1,cap).font=Font(name='맑은 고딕',size=11,bold=True,color=GREEN)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=14)
    row+=1
    src_img=f'/tmp/k23c/{key}.png'
    im=PILImage.open(src_img); W=1150; h=int(im.height*W/im.width)
    im.resize((W,h),PILImage.LANCZOS).save(f'/tmp/k23x/{key}.png')
    xi=XLImage(f'/tmp/k23x/{key}.png'); xi.width=W; xi.height=h
    ws.add_image(xi, f'A{row}')
    row += int(h/19)+3
ws.cell(row,1,'※ 2023학년도는 수치(평균·컷) 표가 공개되지 않아 위 분포표가 원자료 전부입니다. 정밀 비교에는 2024~2026 수치표를 사용하세요.').font=Font(name='맑은 고딕',size=9,color='C00000')
widths(ws,[14]*14)

wb.save(OUT)
print('saved',OUT, os.path.getsize(OUT))
