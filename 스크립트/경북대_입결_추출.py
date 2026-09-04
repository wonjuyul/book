# -*- coding: utf-8 -*-
import openpyxl, re, os, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = '/home/user/book/입시결과'
OUT = '/home/user/book/엑셀/경북대_농업생명_생태환경_6개모집단위_입결.xlsx'

TARGETS = ['응용생명과학부','식물의학과','원예과학과','농업생명과학대학 자율학부','식물자원학과','곤충생명과학과']
def nm(s): return re.sub(r'\s+','',str(s or ''))
NT = {nm(t): t for t in TARGETS}

# 전형 분류: sheet name -> (대분류, 세부구분)
CAT = {
 '1.학생부교과(교과우수)'      : ('학생부교과','일반(교과우수자)'),
 '2.학생부교과(교과지역)'      : ('학생부교과','지역인재'),
 '3.학생부교과(사회통합)'      : ('학생부교과','사회통합(참고)'),
 '4.논술(AAT)'                 : ('논술(AAT)','일반(참고)'),
 '5.학생부종합(일반학생)'      : ('학생부종합','일반학생'),
 '6.학생부종합(지역인재)'      : ('학생부종합','지역인재'),
 '7.학생부종합(농어촌)'        : ('학생부종합','농어촌(참고)'),
 '8.학생부종합(영농창업인재)'  : ('학생부종합','영농창업인재'),
 '9.학생부종합(SW특별)'        : ('학생부종합','SW특별(참고)'),
}
PRIMARY = {'일반(교과우수자)','지역인재','일반학생','영농창업인재'}

CANON = [
 (None,'모집인원','모집인원'),
 (None,'지원인원','지원인원'),
 (None,'경쟁률','경쟁률'),
 (None,'입학인원','입학인원(등록)'),
 (None,'추합최종번호','추합최종번호'),
 (None,'추합인원','추합인원'),
 (None,'인원수','최저기준통과 인원'),
 (None,'실질경쟁률','실질경쟁률(최저적용후)'),
 ('등급','평균','등급 평균'),
 ('등급','표준편차','등급 표준편차'),
 ('등급','0.5','등급 50%컷'),
 ('등급','0.7','등급 70%컷'),
 ('등급','0.85','등급 85%컷'),
 ('반영점수','만점','반영점수 만점'),
 ('반영점수','평균','반영점수 평균'),
 ('반영점수','0.5','반영점수 50%컷'),
 ('반영점수','0.7','반영점수 70%컷'),
 ('반영점수','0.85','반영점수 85%컷'),
]
METRICS = [c[2] for c in CANON]

def ffill(row):
    out=[]; last=None
    for v in row:
        if v is not None and str(v).strip()!='' : last=str(v).strip()
        out.append(last)
    return out

def num(v):
    if v is None: return None
    s=str(v).strip()
    if s in ('','-','·','―','ㅡ'): return None
    s=s.replace(':1','').replace(',','').strip()
    try:
        f=float(s)
        return int(f) if f==int(f) and abs(f)<100000 and '.' not in s else round(f,2)
    except: return s

def extract(path, year):
    wb=openpyxl.load_workbook(path, data_only=True)
    recs=[]
    for sh in wb.sheetnames:
        if sh not in CAT: continue
        ws=wb[sh]
        rows=[list(r) for r in ws.iter_rows(values_only=True)]
        # locate header row (contains 모집인원)
        hi=None
        for i,r in enumerate(rows[:12]):
            if any(nm(c)=='모집인원' for c in r): hi=i; break
        if hi is None: continue
        grp=ffill(rows[hi-1]); sub=[str(c).strip() if c is not None else '' for c in rows[hi]]
        idx={}
        for j,(g,sb) in enumerate(zip(grp,sub)):
            g2=nm(g); s2=nm(sb)
            for cg,cs,canon in CANON:
                if canon in idx: continue
                if s2!=nm(cs): continue
                if cg is None or nm(cg) in g2: idx[canon]=j
        col_dan=0; col_mo=1
        cur_dan=None
        for r in rows[hi+1:]:
            if r[col_dan] is not None and str(r[col_dan]).strip(): cur_dan=re.sub(r'\s+','',str(r[col_dan]))
            key=nm(r[col_mo]) if len(r)>col_mo else ''
            if key not in NT: continue
            rec={'학년도':year,'단과대학':cur_dan,'모집단위':NT[key],
                 '전형구분':CAT[sh][0],'세부전형':CAT[sh][1],'원본시트':sh}
            for canon in METRICS:
                j=idx.get(canon)
                rec[canon]=num(r[j]) if (j is not None and j<len(r)) else None
            recs.append(rec)
    return recs

ALL=[]
for y in (2024,2025,2026):
    p=os.path.join(SRC, f'경북대_{y}_수시정시_지원결과.xlsx')
    got=extract(p,y); ALL+=got
    print(y,'rows',len(got))
json.dump(ALL,open('/tmp/knu_all.json','w'),ensure_ascii=False,indent=1)
print('total',len(ALL))
